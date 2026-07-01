"""Excel「テストケース2」シートから解析評価の正解データを取り込む (Phase 1)。

D-K 列 = エンジニアの障害解析記録 ①〜⑦＋補足。グループ単位 (A〜N) で 1 シナリオ、
⑥(I列, conclusion)=真因が採点の主軸。⑥が空のグループ (D・E の XXXXX プレースホルダ等)
は取り込まない → 実データでは A B C F G H I J K L M N の 12 件が対象。

列マッピング (1-based / openpyxl):
    A(1)=ID  D(4)=①トリガー  E(5)=②初期仮説  F(6)=③辿った経路  G(7)=④決断点
    H(8)=⑤根拠の出所  I(9)=⑥結論(真因)  J(10)=⑦ジュニアの落とし穴  K(11)=補足

使い方:
    python scripts/import_answers.py <xlsx_path> [--sheet テストケース2] [--dry-run]
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import openpyxl

# scripts/ から src/log_analyzer を解決 (editable install が無い場合の保険)
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))
from log_analyzer import storage  # noqa: E402

DEFAULT_SHEET = "テストケース2"

# D-K の列番号 (1-based) → answer_scenarios の本文フィールド
_COL = {
    "trigger": 4,          # D ①トリガー
    "initial_hypothesis": 5,  # E ②初期仮説
    "path": 6,             # F ③辿った経路
    "decision_points": 7,  # G ④決断点
    "evidence_source": 8,  # H ⑤根拠の出所
    "conclusion": 9,       # I ⑥結論(真因) ★採点主軸
    "junior_pitfall": 10,  # J ⑦ジュニアの落とし穴
    "notes": 11,           # K 補足
}

# グループ見出し行: "A.無線AP..." (半角/全角ドット両対応)
_GROUP_HEADER = re.compile(r"^([A-Z])[.．]\s*(.*)$")
# ケース行: "A-01" / "A-03"
_CASE_ROW = re.compile(r"^([A-Z])-\d+")


def _cell(row: tuple, idx1: int) -> str:
    """1-based 列番号のセルを文字列で返す (None/空白は '')。"""
    i = idx1 - 1
    if i < 0 or i >= len(row):
        return ""
    v = row[i]
    return str(v).strip() if v is not None else ""


def parse_scenarios(xlsx_path: str, sheet: str = DEFAULT_SHEET) -> list[dict]:
    """シートを走査し、⑥(conclusion)が非空のグループを解答シナリオとして返す。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"シート '{sheet}' が見つかりません。存在するシート: {wb.sheetnames}")
    ws = wb[sheet]

    scenarios: dict[str, dict] = {}
    cur_letter = ""
    cur_title = ""
    for row in ws.iter_rows(values_only=True):
        c1 = _cell(row, 1)
        if not c1:
            continue
        m_group = _GROUP_HEADER.match(c1)
        if m_group and _cell(row, 2) == "" and _cell(row, 3) == "":
            # グループ見出し行 (ID/ステータス/解析方式が空 = ケース行ではない)
            cur_letter = m_group.group(1)
            cur_title = m_group.group(2).strip()
            continue
        m_case = _CASE_ROW.match(c1)
        if not m_case:
            continue
        letter = m_case.group(1)
        if letter != cur_letter:
            cur_letter = letter  # 見出しを取りこぼしても letter は case から復元
        if letter in scenarios:
            continue  # そのグループは取込済み (先頭 -01 を優先)
        body = {name: _cell(row, col) for name, col in _COL.items()}
        if not body["conclusion"]:
            continue  # ⑥真因が空 → 取込対象外 (D・E プレースホルダ等)
        scenarios[letter] = {
            "scenario_key": letter,
            "title": cur_title,
            **body,
        }
    return [scenarios[k] for k in sorted(scenarios)]


def main() -> int:
    ap = argparse.ArgumentParser(description="Excel テストケース2 から解答シナリオを取込")
    ap.add_argument("xlsx_path", help="Excel ファイルパス")
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    ap.add_argument("--dry-run", action="store_true", help="DB に書かず内容だけ表示")
    args = ap.parse_args()

    # Windows コンソール (cp932) でも日本語/記号を出せるように UTF-8 化
    try:
        sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
    except Exception:  # noqa: BLE001
        pass

    scenarios = parse_scenarios(args.xlsx_path, args.sheet)
    print(f"取込対象: {len(scenarios)} 件 — {', '.join(s['scenario_key'] for s in scenarios)}")
    for s in scenarios:
        concl = s["conclusion"].replace("\n", " ")
        print(f"  [{s['scenario_key']}] {s['title'][:30]} | ⑥: {concl[:60]}…")

    if args.dry_run:
        print("(dry-run: DB へは書き込みません)")
        return 0

    storage.init_db()
    src = Path(args.xlsx_path).name
    for s in scenarios:
        storage.upsert_answer_scenario(
            s["scenario_key"], source_file=src,
            **{k: v for k, v in s.items() if k != "scenario_key"},
        )
    print(f"{len(scenarios)} 件を answer_scenarios に upsert しました。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
